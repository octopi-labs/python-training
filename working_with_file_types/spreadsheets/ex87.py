import os
import openpyxl

current_dir = os.path.dirname(__file__)
filename = os.path.join(current_dir, "test.xlsx")

# Create workbook object
wb = openpyxl.Workbook()

# Get all sheet names in a workbook
# print("Sheets:", wb.get_sheet_names())
print("Sheets:", wb.sheetnames)

# Get the active sheet
sheet = wb.active

# Sheet object
print("Sheet:", sheet)
# Sheet title
print("Sheet Title:", sheet.title)

# Change sheet title
sheet.title = 'CMI Team'

print("Sheets:", wb.sheetnames)

# Save the spreadsheet workbook
wb.save(filename)

# Create a new sheet
wb.create_sheet()

print("Sheets:", wb.sheetnames)

# Create new sheet at the first position
wb.create_sheet(index=0, title='First Sheet')

print("Sheets:", wb.sheetnames)

# Create sheet at indexed position
wb.create_sheet(index=2, title='Middle Sheet')

print("Sheets:", wb.sheetnames)

# get sheet by name
# print(wb.get_sheet_by_name('CMI Team'))
print(wb['CMI Team'])

# Remove sheet
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

print("Sheets:", wb.sheetnames)

sheet = wb['First Sheet']

# Write value to sheet
sheet['A1'] = "Name"

# Get value of a cell
print(sheet['A1'].value)

sheet['B1'] = "Branch"
sheet['C1'] = "Sick Days"
sheet['D1'] = "Salary"

sheet['A2'] = "Rahul Shelke"
sheet['B2'] = "COE"
sheet['C2'] = "2"
sheet['D2'] = "9.0"

wb.save(filename)
