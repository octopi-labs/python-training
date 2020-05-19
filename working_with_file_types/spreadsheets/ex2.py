import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

current_dir = os.path.dirname(__file__)
filename = os.path.join(current_dir, "test.xlsx")

# Load workbook
wb = openpyxl.load_workbook(filename)

print(wb.sheetnames)

sheet = wb['First Sheet']

anotherSheet = wb.active

c = sheet['A2']

print("Row:", c.row)

print("Column:", c.column)

print("Coordinate:", c.coordinate)

print("Value:", c.value)

print(sheet.cell(row=2, column=2))

print(sheet.cell(row=2, column=2).value)

# Get column letter
print("Column Letter:", get_column_letter(1))

# Get index of the string
print("Column index from string:", column_index_from_string('A'))

# Max row
print(sheet.max_row)

# Max column
print(sheet.max_column)

print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
